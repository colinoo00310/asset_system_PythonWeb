"""
资产管理路由
"""

from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from config.database import get_connection
from routes.auth import permission_required
from werkzeug.utils import secure_filename
import os
import uuid

assets_bp = Blueprint('assets', __name__)

def _is_department_user(user):
    return user.get('role') in ('section_chief', 'staff')

def _can_access_department(user, department_id):
    return not _is_department_user(user) or user.get('department_id') == department_id

def _normalize_coordinates(data):
    """校验并规范经纬度；允许都不填，但不允许只填写其中一个。"""
    if 'longitude' not in data and 'latitude' not in data:
        return None

    longitude = data.get('longitude')
    latitude = data.get('latitude')
    longitude_empty = longitude in (None, '')
    latitude_empty = latitude in (None, '')

    if longitude_empty and latitude_empty:
        data['longitude'] = None
        data['latitude'] = None
        return None
    if longitude_empty or latitude_empty:
        return '经度和纬度需要同时填写'

    try:
        longitude = float(longitude)
        latitude = float(latitude)
    except (TypeError, ValueError):
        return '经纬度必须是有效数字'

    if not -180 <= longitude <= 180:
        return '经度必须在 -180～180 之间'
    if not -90 <= latitude <= 90:
        return '纬度必须在 -90～90 之间'

    data['longitude'] = longitude
    data['latitude'] = latitude
    data['coord_type'] = data.get('coord_type') or 'gcj02'
    return None

@assets_bp.route('/', methods=['GET'])
@permission_required('assets', 'view')
def get_assets():
    """获取资产列表"""
    conn = get_connection()
    c = conn.cursor()
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    department_id = request.args.get('department_id', type=int)
    
    where_clauses = []
    params = []
    current_user = request.current_user

    if _is_department_user(current_user):
        where_clauses.append("a.department_id = ?")
        params.append(current_user.get('department_id'))
    
    if search:
        keyword = f'%{search}%'
        where_clauses.append("""(
            a.name LIKE ? OR a.asset_number LIKE ? OR a.location LIKE ?
            OR a.responsible_person LIKE ? OR d.name LIKE ?
            OR a.management_type LIKE ? OR a.category LIKE ?
            OR a.status LIKE ? OR a.lease_status LIKE ?
        )""")
        params.extend([keyword] * 9)
    
    if category:
        management_types = ('租赁管理', '自主管理', '托管管理')
        where_clauses.append("a.management_type = ?" if category in management_types else "a.category = ?")
        params.append(category)
    
    if department_id:
        where_clauses.append("a.department_id = ?")
        params.append(department_id)
    
    where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
    
    c.execute(f"""SELECT COUNT(*)
                  FROM assets a
                  LEFT JOIN departments d ON a.department_id = d.id
                  WHERE {where_sql}""", params)
    total = c.fetchone()[0]
    
    offset = (page - 1) * per_page
    c.execute(f"""SELECT a.*, d.name as department_name
                  FROM assets a
                  LEFT JOIN departments d ON a.department_id = d.id
                  WHERE {where_sql}
                  ORDER BY a.id ASC LIMIT ? OFFSET ?""",
              params + [per_page, offset])
    raw_assets = c.fetchall()

    c2 = get_connection().cursor()
    c2.execute("PRAGMA table_info(assets)")
    columns = [col[1] for col in c2.fetchall()]
    columns.append('department_name')
    conn.close()

    result = [dict(zip(columns, row)) for row in raw_assets]
    for asset in result:
        if asset.get('management_type') == '租赁管理' and not str(asset.get('lease_status') or '').strip():
            asset['lease_status'] = '租赁中'
    
    return jsonify({
        'assets': result,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page
    })

@assets_bp.route('/<int:asset_id>', methods=['GET'])
@permission_required('assets', 'view')
def get_asset(asset_id):
    """获取单个资产详情"""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT a.*, d.name as department_name FROM assets a LEFT JOIN departments d ON a.department_id = d.id WHERE a.id=?", (asset_id,))
    raw_asset = c.fetchone()

    if raw_asset:
        c.execute('PRAGMA table_info(assets)')
        base_columns = [col[1] for col in c.fetchall()]
        department_index = base_columns.index('department_id')
        if not _can_access_department(request.current_user, raw_asset[department_index]):
            conn.close()
            return jsonify({'error': '资产不存在'}), 404

    c2 = get_connection().cursor()
    c2.execute("PRAGMA table_info(assets)")
    columns = [col[1] for col in c2.fetchall()]
    columns.append('department_name')
    conn.close()

    if not raw_asset:
        return jsonify({'error': '资产不存在'}), 404

    result = dict(zip(columns, raw_asset))
    if result.get('management_type') == '租赁管理' and not str(result.get('lease_status') or '').strip():
        result['lease_status'] = '租赁中'
    return jsonify(result)

@assets_bp.route('/', methods=['POST'])
@permission_required('assets', 'add')
def create_asset():
    """创建资产"""
    data = request.get_json()
    current_user = request.current_user

    coordinate_error = _normalize_coordinates(data)
    if coordinate_error:
        return jsonify({'error': coordinate_error}), 400

    if _is_department_user(current_user):
        if not current_user.get('department_id'):
            return jsonify({'error': '当前账号未分配部门，无法新增资产'}), 403
        data['department_id'] = current_user['department_id']
    
    required_fields = ['name', 'category', 'asset_number', 'responsible_person', 'location']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'{field} 不能为空'}), 400
    
    conn = get_connection()
    c = conn.cursor()
    
    if _is_department_user(current_user):
        data['department_id'] = current_user['department_id']
    
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    fields = ['name', 'category', 'management_type', 'asset_number', 'quantity',
              'model', 'purchase_date', 'market_value', 'responsible_person', 'location',
              'status', 'lease_start_date', 'lease_end_date', 'lease_reminder_days', 'lease_status',
              'tenant_name', 'tenant_contact', 'notes', 'department_id',
              'certificate_status', 'property_unit', 'building_area',
              'trusteeship_contract_type', 'trusteeship_contract_amount',
              'trusteeship_counterparty', 'trusteeship_contract_number',
              'trusteeship_start_date', 'trusteeship_end_date', 'trusteeship_sign_date',
              'trusteeship_is_archived', 'tenant_nature', 'tenant_purpose',
              'rent_amount', 'rent_payment_method', 'bidding_situation',
              'longitude', 'latitude', 'coord_type']
    
    placeholders = []
    params = []
    
    for field in fields:
        if field in data:
            placeholders.append(field)
            params.append(data[field])
    
    placeholders.extend(['created_by', 'created_at'])
    params.extend([current_user['username'], current_time])
    
    try:
        c.execute(f"""INSERT INTO assets ({', '.join(placeholders)})
                      VALUES ({', '.join(['?'] * len(params))})""", params)
        asset_id = c.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'message': '资产创建成功', 'id': asset_id}), 201
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400

@assets_bp.route('/<int:asset_id>', methods=['PUT'])
@permission_required('assets', 'edit')
def update_asset(asset_id):
    """更新资产"""
    data = request.get_json()

    coordinate_error = _normalize_coordinates(data)
    if coordinate_error:
        return jsonify({'error': coordinate_error}), 400
    
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("SELECT id, department_id FROM assets WHERE id=?", (asset_id,))
    existing_asset = c.fetchone()
    if not existing_asset:
        conn.close()
        return jsonify({'error': '资产不存在'}), 404
    if not _can_access_department(request.current_user, existing_asset[1]):
        conn.close()
        return jsonify({'error': '您只能编辑所属部门的资产'}), 403
    if _is_department_user(request.current_user):
        data['department_id'] = request.current_user['department_id']
    
    if _is_department_user(request.current_user):
        data['department_id'] = request.current_user['department_id']
    
    update_fields = []
    params = []
    
    fields = ['name', 'category', 'management_type', 'asset_number', 'quantity',
              'model', 'purchase_date', 'market_value', 'responsible_person', 'location',
              'status', 'lease_start_date', 'lease_end_date', 'lease_reminder_days', 'lease_status',
              'tenant_name', 'tenant_contact', 'notes', 'department_id',
              'certificate_status', 'property_unit', 'building_area',
              'trusteeship_contract_type', 'trusteeship_contract_amount',
              'trusteeship_counterparty', 'trusteeship_contract_number',
              'trusteeship_start_date', 'trusteeship_end_date', 'trusteeship_sign_date',
              'trusteeship_is_archived', 'tenant_nature', 'tenant_purpose',
              'rent_amount', 'rent_payment_method', 'bidding_situation',
              'longitude', 'latitude', 'coord_type']
    
    for field in fields:
        if field in data:
            update_fields.append(f"{field} = ?")
            params.append(data[field])
    
    if not update_fields:
        conn.close()
        return jsonify({'error': '没有要更新的字段'}), 400
    
    params.append(asset_id)
    
    try:
        c.execute(f"UPDATE assets SET {', '.join(update_fields)} WHERE id=?", params)
        conn.commit()
        conn.close()
        
        return jsonify({'message': '资产更新成功'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400

@assets_bp.route('/<int:asset_id>', methods=['DELETE'])
@permission_required('assets', 'delete')
def delete_asset(asset_id):
    """删除资产"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("SELECT id, department_id FROM assets WHERE id=?", (asset_id,))
    existing_asset = c.fetchone()
    if not existing_asset:
        conn.close()
        return jsonify({'error': '资产不存在'}), 404
    if not _can_access_department(request.current_user, existing_asset[1]):
        conn.close()
        return jsonify({'error': '您只能删除所属部门的资产'}), 403
    
    try:
        c.execute("DELETE FROM assets WHERE id=?", (asset_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'message': '资产删除成功'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400

@assets_bp.route('/<int:asset_id>/images', methods=['POST'])
@permission_required('assets', 'edit')
def upload_asset_images(asset_id):
    """同步资产图片：保留指定旧图、加入新图、删除已移除图片，最多三张。"""
    conn = get_connection(); c = conn.cursor()
    c.execute('SELECT department_id, image_path1, image_path2, image_path3 FROM assets WHERE id=?', (asset_id,))
    asset = c.fetchone()
    if not asset: conn.close(); return jsonify({'error': '资产不存在'}), 404
    if not _can_access_department(request.current_user, asset[0]):
        conn.close(); return jsonify({'error': '您只能编辑所属部门的资产'}), 403
    files = [item for item in request.files.getlist('images') if item and item.filename]
    original_paths = [path for path in asset[1:] if path]
    requested_keep = request.form.getlist('keep_images')
    keep_paths = [path for path in requested_keep if path in original_paths]
    if len(keep_paths) + len(files) > 3: conn.close(); return jsonify({'error': '最多保留或上传3张图片'}), 400
    allowed = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}
    images_dir = os.environ.get('ASSET_IMAGES_DIR') or os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'assets', 'images')
    os.makedirs(images_dir, exist_ok=True); saved = []
    try:
        for index, upload in enumerate(files):
            # 中文文件名经过 secure_filename 后可能只剩 "png"，扩展名应从原始文件名提取。
            original_name = upload.filename or ''
            extension = os.path.splitext(original_name)[1].lower()
            if extension not in allowed: raise ValueError(f'不支持的图片格式：{upload.filename}')
            filename = f'{asset_id}_{index}_{uuid.uuid4().hex[:10]}{extension}'
            upload.save(os.path.join(images_dir, filename)); saved.append(f'assets/images/{filename}')
        final_paths = keep_paths + saved
        values = final_paths + [None] * (3 - len(final_paths))
        c.execute('UPDATE assets SET image_path1=?, image_path2=?, image_path3=? WHERE id=?', (*values, asset_id))
        conn.commit(); conn.close()
        # 数据库提交后再清理用户明确移除的、位于资产图片目录内的文件。
        for old_path in original_paths:
            if old_path in keep_paths:
                continue
            old_file = os.path.abspath(os.path.join(images_dir, os.path.basename(old_path)))
            if os.path.dirname(old_file) == os.path.abspath(images_dir):
                try: os.remove(old_file)
                except OSError: pass
        return jsonify({'message': '图片保存成功', 'images': final_paths})
    except Exception as exc:
        conn.rollback(); conn.close()
        for path in saved:
            try: os.remove(os.path.join(images_dir, os.path.basename(path)))
            except OSError: pass
        return jsonify({'error': str(exc)}), 400

@assets_bp.route('/categories', methods=['GET'])
@permission_required('assets', 'view')
def get_categories():
    """获取资产分类列表"""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT category FROM assets ORDER BY category")
    categories = [row[0] for row in c.fetchall() if row[0]]
    for management_type in ('租赁管理', '自主管理', '托管管理'):
        if management_type not in categories:
            categories.append(management_type)
    conn.close()
    
    return jsonify({'categories': categories})

@assets_bp.route('/export', methods=['GET'])
@permission_required('assets', 'export')
def export_assets_excel():
    """导出资产列表为 Excel（与原系统格式一致）"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from io import BytesIO

        conn = get_connection()
        c = conn.cursor()
        query = """SELECT a.id, a.name, a.management_type, a.category, a.asset_number,
                            a.quantity, a.purchase_date, a.market_value,
                            a.responsible_person, a.location, a.status,
                            a.lease_start_date, a.lease_end_date, a.tenant_name, a.tenant_contact,
                            a.tenant_nature, a.tenant_purpose, a.rent_amount, a.rent_payment_method, a.bidding_situation,
                            a.certificate_status, a.property_unit, a.building_area,
                            a.trusteeship_contract_type, a.trusteeship_contract_amount, a.trusteeship_counterparty,
                            a.trusteeship_contract_number, a.trusteeship_start_date, a.trusteeship_end_date, a.trusteeship_sign_date,
                            a.trusteeship_is_archived,
                            COALESCE(u.full_name, a.created_by) as display_name,
                            d.name as department_name,
                            a.notes
                     FROM assets a
                     LEFT JOIN users u ON a.created_by = u.username
                     LEFT JOIN departments d ON a.department_id = d.id"""
        params = []
        current_user = request.current_user
        if current_user['role'] in ('section_chief', 'staff'):
            query += ' WHERE a.department_id = ?'
            params.append(current_user.get('department_id'))
        query += ' ORDER BY a.id'
        c.execute(query, params)
        rows = c.fetchall()
        conn.close()

        headers = [
            "ID", "名称", "管理方式", "资产分类", "编号", "数量",
            "购置日期", "市场价值", "责任人", "地址", "状态",
            "租赁开始日期", "租赁结束日期", "承租方", "承租方联系方式",
            "承租方性质", "承租方用途", "租金金额", "租金交付方式", "公开招拍租情况",
            "产证情况", "产权单位", "建筑面积（平方米）",
            "托管合同类型", "托管合同金额", "托管合同相对方",
            "托管合同编号", "托管合同开始日期", "托管合同结束日期", "托管签署日期",
            "托管是否归档",
            "录入人", "所属部门", "备注"
        ]

        if not rows or len(rows[0]) != len(headers):
            return jsonify({'error': '数据字段与导出格式不匹配，请联系管理员'}), 500

        wb = Workbook()
        ws = wb.active
        ws.title = "资产清单"
        ws.append(headers)

        for display_id, row in enumerate(rows, 1):
            formatted_row = []
            formatted_row.append(display_id)

            management_type = row[2] if len(row) > 2 else ""

            for i in range(1, len(row)):
                value = row[i]
                is_trusteeship_field = 24 <= i <= 30

                if is_trusteeship_field and management_type != "托管管理":
                    formatted_row.append("")
                elif i == 5:
                    try:
                        formatted_row.append(int(float(value)) if value is not None else "")
                    except (ValueError, TypeError):
                        formatted_row.append("")
                elif i == 7 or i == 17 or i == 24:
                    try:
                        if value is not None and value != "" and value != 0:
                            formatted_row.append(float(value))
                        else:
                            formatted_row.append("")
                    except Exception:
                        formatted_row.append(str(value).strip() if value is not None else "")
                elif i in [6, 11, 12, 27, 28, 29]:
                    if value is None or str(value).strip() == "" or str(value).strip() == "-":
                        formatted_row.append("")
                    else:
                        try:
                            date_text = str(value).strip()[:10]
                            datetime.strptime(date_text, "%Y-%m-%d")
                            formatted_row.append(date_text)
                        except Exception:
                            formatted_row.append("")
                else:
                    formatted_row.append(str(value).strip() if value is not None else "")

            ws.append(formatted_row)

        col_widths = [10, 25, 12, 12, 15, 8, 12, 15, 12, 25, 10, 15, 15, 15, 15, 12, 20, 15, 15, 20, 20, 20, 20, 15, 15, 20, 15, 18, 18, 16, 12, 12, 15, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 30

        header_font = Font(bold=True, name='微软雅黑', size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        date_columns = ['G', 'L', 'M']
        for col in date_columns:
            for cell in ws[col][1:]:
                if cell.value and str(cell.value).strip():
                    try:
                        datetime.strptime(str(cell.value).strip(), "%Y-%m-%d")
                        cell.number_format = 'yyyy-mm-dd'
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception:
                        cell.number_format = '@'
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws['H'][1:]:
            if cell.value is not None and cell.value != "":
                try:
                    cell.value = float(cell.value)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                except (ValueError, TypeError):
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws['R'][1:]:
            if cell.value is not None and cell.value != "":
                try:
                    cell.value = float(cell.value)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                except (ValueError, TypeError):
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws['Y'][1:]:
            if cell.value is not None and cell.value != "":
                try:
                    cell.value = float(cell.value)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                except (ValueError, TypeError):
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        trusteeship_date_columns = ['AA', 'AB', 'AC']
        for col in trusteeship_date_columns:
            for cell in ws[col][1:]:
                if cell.value and str(cell.value).strip():
                    try:
                        datetime.strptime(str(cell.value).strip(), "%Y-%m-%d")
                        cell.number_format = 'yyyy-mm-dd'
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception:
                        cell.number_format = '@'
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.alignment.horizontal is None:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        max_row = ws.max_row
        max_col = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = thin_border

        ws.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"资产清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f'导出Excel失败: {str(e)}')
        return jsonify({'error': f'导出失败：{str(e)}'}), 500

@assets_bp.route('/import-template', methods=['GET'])
@permission_required('assets', 'import')
def download_asset_import_template():
    """下载与单机版导入格式一致的资产模板。"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    headers = [
        "ID", "名称", "管理方式", "资产分类", "编号", "数量", "购置日期", "市场价值",
        "责任人", "地址", "状态", "租赁开始日期", "租赁结束日期", "承租方", "承租方联系方式",
        "承租方性质", "承租方用途", "租金金额", "租金交付方式", "公开招拍租情况", "产证情况",
        "产权单位", "建筑面积（平方米）", "托管合同类型", "托管合同金额", "托管合同相对方",
        "托管合同编号", "托管合同开始日期", "托管合同结束日期", "托管签署日期", "托管是否归档",
        "录入人", "所属部门", "备注"
    ]
    example = [1, '示例资产（导入前请删除本行）', '自主管理', '房屋资产', '示例编号001', 1,
               '2026-01-01', 100000, '张三', '杭州市示例地址', '正常使用', '', '', '', '', '', '', '',
               '', '', '有证', '示例产权单位', 100, '', '', '', '', '', '', '', '否', 'admin', '默认部门', '示例']
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '资产清单'; ws.append(headers); ws.append(example)
    for cell in ws[1]: cell.font = Font(bold=True, name='微软雅黑'); cell.alignment = Alignment(horizontal='center')
    for column_index, header in enumerate(headers, 1):
        example_text = str(example[column_index - 1] or '')
        width = min(max(len(str(header)) * 2 + 4, len(example_text) + 4, 12), 36)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = width
    ws.freeze_panes = 'A2'
    output = BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name='资产导入模板.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@assets_bp.route('/import', methods=['POST'])
@permission_required('assets', 'import')
def import_assets_excel():
    """导入 Excel 资产列表（与原系统格式兼容）"""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '请上传文件'}), 400

    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return jsonify({'error': 'Excel 内容为空'}), 400

    header = [str(x).strip() if x else '' for x in rows[0]]
    expected_headers = [
        "ID", "名称", "管理方式", "资产分类", "编号", "数量",
        "购置日期", "市场价值", "责任人", "地址", "状态",
        "租赁开始日期", "租赁结束日期", "承租方", "承租方联系方式",
        "承租方性质", "承租方用途", "租金金额", "租金交付方式", "公开招拍租情况",
        "产证情况", "产权单位", "建筑面积（平方米）",
        "托管合同类型", "托管合同金额", "托管合同相对方",
        "托管合同编号", "托管合同开始日期", "托管合同结束日期", "托管签署日期",
        "托管是否归档",
        "录入人", "所属部门", "备注"
    ]

    if header != expected_headers:
        return jsonify({'error': f'Excel文件格式不正确，请使用本系统导出的Excel文件。当前标题行：{header}'}), 400

    imported_data = []
    for row in rows[1:]:
        if not any(row):
            continue
        imported_data.append(list(row))

    if not imported_data:
        return jsonify({'error': 'Excel文件中没有有效数据'}), 400

    conn = get_connection()
    c = conn.cursor()

    c.execute("SELECT id, username, full_name FROM users")
    users = c.fetchall()
    username_to_fullname = {}
    fullname_to_username = {}
    for user_id, username, full_name in users:
        username_to_fullname[username] = full_name or username
        fullname_to_username[full_name] = username
        if username == "admin":
            fullname_to_username["系统管理员"] = "admin"
        if full_name == "系统管理员":
            fullname_to_username["系统管理员"] = "admin"

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    current_user = request.current_user
    current_username = current_user['username']

    imported_count = 0
    duplicate_count = 0
    errors = []

    def clean_date(date_value):
        if not date_value or str(date_value).strip() == "":
            return None
        date_str = str(date_value).strip()
        if ' ' in date_str:
            date_str = date_str.split(' ')[0]
        try:
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"]:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue
            return date_str
        except Exception:
            return None

    for i, row in enumerate(imported_data, start=2):
        try:
            if len(row) < 34:
                errors.append(f"第{i}行: 数据列不足")
                continue

            asset_name = row[1] if len(row) > 1 else ""
            management_type = row[2] if len(row) > 2 else ""
            category = row[3] if len(row) > 3 else ""
            asset_number = row[4] if len(row) > 4 else ""
            quantity_str = row[5] if len(row) > 5 else ""
            try:
                quantity = int(float(quantity_str)) if quantity_str not in [None, ""] else 1
            except (ValueError, TypeError):
                quantity = 1
                errors.append(f"第{i}行: 数量格式错误，已设为1")

            purchase_date_db = clean_date(row[6])
            market_value_db = None
            market_value_raw = row[7]
            if market_value_raw and market_value_raw != "" and market_value_raw != 0:
                try:
                    market_value_db = float(str(market_value_raw).replace('¥', '').replace(',', '').strip())
                except (ValueError, TypeError):
                    market_value_db = None
                    errors.append(f"第{i}行: 市场价值格式错误，已设为空")

            responsible_person = row[8] if len(row) > 8 else None
            location = row[9] if len(row) > 9 else ""
            if not location or str(location).strip() == "":
                location = "未填写地址"
            status = row[10] if len(row) > 10 else "正常使用"
            lease_start_date_db = clean_date(row[11])
            lease_end_date_db = clean_date(row[12])
            tenant_name = row[13] if len(row) > 13 else None
            tenant_contact = row[14] if len(row) > 14 else None
            tenant_nature = row[15] if len(row) > 15 else None
            tenant_purpose = row[16] if len(row) > 16 else None
            rent_amount_db = None
            rent_amount_raw = row[17]
            if rent_amount_raw and rent_amount_raw != "" and rent_amount_raw != 0:
                try:
                    rent_amount_db = float(str(rent_amount_raw).replace('¥', '').replace(',', '').strip())
                except (ValueError, TypeError):
                    rent_amount_db = None
                    errors.append(f"第{i}行: 租金金额格式错误，已设为空")
            rent_payment_method = row[18] if len(row) > 18 else None
            bidding_situation = row[19] if len(row) > 19 else None
            certificate_status = row[20] if len(row) > 20 else None
            property_unit = row[21] if len(row) > 21 else None
            building_area = row[22] if len(row) > 22 else None
            trusteeship_contract_type = row[23] if len(row) > 23 else None
            trusteeship_contract_amount_db = row[24] if len(row) > 24 else None
            if trusteeship_contract_amount_db is not None and str(trusteeship_contract_amount_db).strip() == "":
                trusteeship_contract_amount_db = None
            trusteeship_counterparty = row[25] if len(row) > 25 else None
            trusteeship_contract_number = row[26] if len(row) > 26 else None
            trusteeship_start_date_db = clean_date(row[27])
            trusteeship_end_date_db = clean_date(row[28])
            trusteeship_sign_date_db = clean_date(row[29])
            trusteeship_is_archived = row[30] if len(row) > 30 else "否"
            created_by_excel = row[31] if len(row) > 31 else ""
            created_by_db = current_username

            if created_by_excel and str(created_by_excel).strip():
                excel_input = str(created_by_excel).strip()
                if excel_input in username_to_fullname:
                    created_by_db = excel_input
                elif excel_input in fullname_to_username:
                    created_by_db = fullname_to_username[excel_input]
                elif excel_input == "系统管理员":
                    created_by_db = "admin"
                else:
                    errors.append(f"第{i}行: 录入人 '{excel_input}' 未找到对应账户，已使用 '{current_username}'")

            department_name = row[32] if len(row) > 32 else None
            notes = row[33] if len(row) > 33 else None

            if not asset_name or not asset_number:
                errors.append(f"第{i}行: 名称或编号不能为空")
                continue

            c.execute("SELECT 1 FROM assets WHERE asset_number=?", (asset_number,))
            if c.fetchone():
                duplicate_count += 1
                errors.append(f"第{i}行: 编号 '{asset_number}' 已存在，已跳过")
                continue

            department_id = None
            if department_name and str(department_name).strip():
                c.execute("SELECT id FROM departments WHERE name=?", (str(department_name).strip(),))
                dept_result = c.fetchone()
                if dept_result:
                    department_id = dept_result[0]
                else:
                    errors.append(f"第{i}行: 部门 '{department_name}' 不存在，将不设置部门")

            if department_id is None and current_user['role'] in ('section_chief', 'staff'):
                department_id = current_user.get('department_id')

            if current_user['role'] in ('section_chief', 'staff') and department_id != current_user.get('department_id'):
                errors.append(f"第{i}行: 您只能导入自己部门的资产")
                continue

            if management_type not in ["自主管理", "租赁管理", "托管管理"]:
                errors.append(f"第{i}行: 管理方式 '{management_type}' 无效，应为：自主管理/租赁管理/托管管理")
                continue

            if management_type == "租赁管理" and not lease_end_date_db:
                errors.append(f"第{i}行: 租赁管理资产必须填写租赁结束日期")
                continue

            if management_type == "托管管理":
                if trusteeship_end_date_db:
                    try:
                        datetime.strptime(trusteeship_end_date_db, "%Y-%m-%d")
                    except ValueError:
                        errors.append(f"第{i}行: 托管合同结束日期格式不正确")
                        continue
                if trusteeship_start_date_db:
                    try:
                        datetime.strptime(trusteeship_start_date_db, "%Y-%m-%d")
                    except ValueError:
                        errors.append(f"第{i}行: 托管合同开始日期格式不正确")
                        continue

            params = [
                asset_name, category, management_type, asset_number, quantity,
                None,
                purchase_date_db, market_value_db, responsible_person, location, status,
                lease_start_date_db, lease_end_date_db, 30, '租赁中' if management_type == '租赁管理' else None,
                tenant_name, tenant_contact, None, None, None, notes,
                created_by_db, current_time, department_id,
                certificate_status, property_unit, building_area,
                trusteeship_contract_type, trusteeship_contract_amount_db,
                trusteeship_counterparty, trusteeship_contract_number,
                trusteeship_start_date_db, trusteeship_end_date_db, trusteeship_sign_date_db,
                trusteeship_is_archived, tenant_nature, tenant_purpose,
                rent_amount_db, rent_payment_method, bidding_situation
            ]

            c.execute('''INSERT INTO assets
                        (name, category, management_type, asset_number, quantity, model,
                         purchase_date, market_value, responsible_person,
                         location, status, lease_start_date, lease_end_date,
                         lease_reminder_days, lease_status, tenant_name, tenant_contact,
                         image_path1, image_path2, image_path3, notes,
                         created_by, created_at, department_id,
                         certificate_status, property_unit, building_area,
                         trusteeship_contract_type, trusteeship_contract_amount,
                         trusteeship_counterparty, trusteeship_contract_number,
                         trusteeship_start_date, trusteeship_end_date, trusteeship_sign_date,
                         trusteeship_is_archived, tenant_nature, tenant_purpose,
                         rent_amount, rent_payment_method, bidding_situation)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        params)
            imported_count += 1
        except Exception as e:
            errors.append(f"第{i}行: {str(e)}")
            import traceback
            traceback.print_exc()
            continue

    conn.commit()
    conn.close()

    result = {'imported': imported_count}
    if duplicate_count:
        result['duplicate'] = duplicate_count
    if errors:
        result['errors'] = errors[:50]
    return jsonify(result)

@assets_bp.route('/reminders', methods=['GET'])
@permission_required('assets', 'view')
def get_reminders():
    """提醒列表：租赁到期、托管到期"""
    try:
        reminder_type = request.args.get('type', 'all')
        conn = get_connection()
        c = conn.cursor()
        scoped = _is_department_user(request.current_user)
        scope_sql = ' AND a.department_id=?' if scoped else ''
        scope_params = (request.current_user.get('department_id'),) if scoped else ()

        results = {}
        if reminder_type in ('all', 'lease'):
            c.execute("""SELECT a.id, a.name, a.lease_end_date, a.lease_status, a.tenant_name, a.tenant_contact
                         FROM assets a
                         WHERE a.management_type='租赁管理' AND a.lease_end_date IS NOT NULL AND a.lease_end_date <> ''""" + scope_sql, scope_params)
            lease_rows = c.fetchall()
            lease_items = []
            for row in lease_rows:
                aid, name, end_date, lease_status, tenant, contact = row
                if not end_date:
                    continue
                try:
                    days = (datetime.strptime(end_date, '%Y-%m-%d') - datetime.now()).days
                except Exception:
                    days = None
                lease_items.append({
                    'id': aid, 'name': name, 'end_date': end_date,
                    'status': lease_status or '租赁中', 'tenant_name': tenant, 'contact': contact, 'days': days
                })
            results['lease'] = lease_items

        if reminder_type in ('all', 'trusteeship'):
            c.execute("""SELECT a.id, a.name, a.trusteeship_end_date, a.trusteeship_counterparty, a.trusteeship_contract_number
                         FROM assets a
                         WHERE a.management_type='托管管理' AND a.trusteeship_end_date IS NOT NULL AND a.trusteeship_end_date <> ''""" + scope_sql, scope_params)
            trusteeship_rows = c.fetchall()
            trusteeship_items = []
            for row in trusteeship_rows:
                aid, name, end_date, counterparty, contract_number = row
                if not end_date:
                    continue
                try:
                    days = (datetime.strptime(end_date, '%Y-%m-%d') - datetime.now()).days
                except Exception:
                    days = None
                trusteeship_items.append({
                    'id': aid, 'name': name, 'end_date': end_date,
                    'counterparty': counterparty, 'contract_number': contract_number, 'days': days
                })
            results['trusteeship'] = trusteeship_items

        conn.close()
        return jsonify(results)
    except Exception as e:
        print(f'获取提醒失败: {str(e)}')
        return jsonify({'error': f'获取提醒失败：{str(e)}'}), 500

@assets_bp.route('/stats', methods=['GET'])
@permission_required('assets', 'view')
def get_stats():
    """获取资产统计"""
    conn = get_connection()
    c = conn.cursor()
    where = ' WHERE department_id=?' if _is_department_user(request.current_user) else ''
    params = (request.current_user.get('department_id'),) if where else ()

    c.execute("SELECT COUNT(*) FROM assets" + where, params)
    total = c.fetchone()[0]
    
    c.execute("SELECT status, COUNT(*) FROM assets" + where + " GROUP BY status", params)
    by_status = {row[0] or '未设置': row[1] for row in c.fetchall()}
    
    c.execute("SELECT category, COUNT(*) FROM assets" + where + " GROUP BY category", params)
    by_category = {row[0]: row[1] for row in c.fetchall()}
    
    value_where = ' WHERE market_value IS NOT NULL' + (' AND department_id=?' if where else '')
    c.execute("SELECT SUM(market_value) FROM assets" + value_where, params)
    total_value = c.fetchone()[0] or 0
    
    # 管理类型统计
    c.execute("SELECT management_type, COUNT(*) FROM assets" + where + " GROUP BY management_type", params)
    by_management_type = {}
    for row in c.fetchall():
        management_type = row[0] if row[0] else '未设置'
        by_management_type[management_type] = row[1]
    
    # 用户数量
    c.execute("SELECT COUNT(*) FROM users")
    user_count = c.fetchone()[0]
    
    # 部门数量（只统计有员工的部门）
    c.execute("SELECT COUNT(DISTINCT d.id) FROM departments d INNER JOIN employees e ON d.id = e.department_id")
    department_count = c.fetchone()[0]
    
    conn.close()
    
    return jsonify({
        'total': total,
        'by_status': by_status,
        'by_category': by_category,
        'total_value': total_value,
        'by_management_type': by_management_type,
        'user_count': user_count,
        'department_count': department_count
    })
