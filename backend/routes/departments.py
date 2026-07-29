"""
部门管理路由
"""

from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from config.database import get_connection
from routes.auth import permission_required

departments_bp = Blueprint('departments', __name__)

DEPARTMENT_EMPLOYEE_HEADERS = [
    "部门名称", "负责人", "联系方式", "部门地址",
    "部门内员工序号", "员工姓名", "员工职位", "员工联系方式"
]

OLD_DEPARTMENT_EMPLOYEE_HEADERS = [
    "部门名称", "部门地址", "部门联系方式", "部门负责人",
    "部门内员工序号", "员工姓名", "员工职位", "员工联系方式"
]

@departments_bp.route('/', methods=['GET'])
@permission_required('departments', 'view')
def get_departments():
    """获取部门列表"""
    conn = get_connection()
    c = conn.cursor()
    
    search = request.args.get('search', '')
    
    if search:
        c.execute("""SELECT d.*, COUNT(e.id) as employee_count 
                     FROM departments d
                     LEFT JOIN employees e ON d.id = e.department_id
                     WHERE d.name LIKE ? OR d.address LIKE ? OR d.manager LIKE ?
                     GROUP BY d.id
                     ORDER BY d.id""",
                  (f'%{search}%', f'%{search}%', f'%{search}%'))
    else:
        c.execute("""SELECT d.*, COUNT(e.id) as employee_count 
                     FROM departments d
                     LEFT JOIN employees e ON d.id = e.department_id
                     GROUP BY d.id
                     ORDER BY d.id""")
    
    departments = c.fetchall()
    conn.close()
    
    c = get_connection().cursor()
    c.execute("PRAGMA table_info(departments)")
    columns = [col[1] for col in c.fetchall()]
    columns.append('employee_count')
    conn.close()
    
    result = []
    for dept in departments:
        item = dict(zip(columns, dept))
        result.append(item)
    
    return jsonify({'departments': result})

@departments_bp.route('/<int:dept_id>', methods=['GET'])
@permission_required('departments', 'view')
def get_department(dept_id):
    """获取单个部门详情"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("""SELECT d.*, COUNT(e.id) as employee_count 
                 FROM departments d
                 LEFT JOIN employees e ON d.id = e.department_id
                 WHERE d.id=?
                 GROUP BY d.id""", (dept_id,))
    dept = c.fetchone()
    conn.close()
    
    if not dept:
        return jsonify({'error': '部门不存在'}), 404
    
    c = get_connection().cursor()
    c.execute("PRAGMA table_info(departments)")
    columns = [col[1] for col in c.fetchall()]
    columns.append('employee_count')
    conn.close()
    
    return jsonify(dict(zip(columns, dept)))

@departments_bp.route('/', methods=['POST'])
@permission_required('departments', 'add')
def create_department():
    """创建部门"""
    data = request.get_json()
    current_user = request.current_user
    
    if current_user['role'] != 'admin':
        return jsonify({'error': '只有管理员可以创建部门'}), 403
    
    name = data.get('name', '').strip()
    manager = data.get('manager', '').strip()
    address = data.get('address', '').strip()
    contact = data.get('contact', '').strip()
    
    if not name or not manager or not address:
        return jsonify({'error': '部门名称、负责人、地址不能为空'}), 400
    
    conn = get_connection()
    c = conn.cursor()
    
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    try:
        c.execute("""INSERT INTO departments (name, manager, contact, address, created_at)
                     VALUES (?, ?, ?, ?, ?)""",
                  (name, manager, contact, address, current_time))
        dept_id = c.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'message': '部门创建成功', 'id': dept_id}), 201
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400

@departments_bp.route('/<int:dept_id>', methods=['PUT'])
@permission_required('departments', 'edit')
def update_department(dept_id):
    """更新部门"""
    data = request.get_json()
    current_user = request.current_user
    
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("SELECT id, manager FROM departments WHERE id=?", (dept_id,))
    dept = c.fetchone()
    if not dept:
        conn.close()
        return jsonify({'error': '部门不存在'}), 404
    
    if current_user['role'] == 'section_chief':
        if current_user.get('department_id') != dept_id:
            conn.close()
            return jsonify({'error': '您只能编辑自己所在的部门'}), 403
        update_fields = []
        params = []
        if 'contact' in data:
            update_fields.append("contact = ?")
            params.append(data['contact'])
        if 'address' in data:
            update_fields.append("address = ?")
            params.append(data['address'])
    elif current_user['role'] == 'admin':
        update_fields = []
        params = []
        for field in ['name', 'manager', 'contact', 'address']:
            if field in data:
                update_fields.append(f"{field} = ?")
                params.append(data[field])
    else:
        conn.close()
        return jsonify({'error': '权限不足'}), 403
    
    if not update_fields:
        conn.close()
        return jsonify({'error': '没有要更新的字段'}), 400
    
    params.append(dept_id)
    
    try:
        c.execute(f"UPDATE departments SET {', '.join(update_fields)} WHERE id=?", params)
        conn.commit()
        conn.close()
        
        return jsonify({'message': '部门更新成功'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400

@departments_bp.route('/<int:dept_id>', methods=['DELETE'])
@permission_required('departments', 'delete')
def delete_department(dept_id):
    """删除部门"""
    current_user = request.current_user
    
    if current_user['role'] != 'admin':
        return jsonify({'error': '只有管理员可以删除部门'}), 403
    
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("SELECT COUNT(*) FROM employees WHERE department_id=?", (dept_id,))
    if c.fetchone()[0] > 0:
        conn.close()
        return jsonify({'error': '该部门下有员工，无法删除'}), 400
    
    try:
        c.execute("DELETE FROM departments WHERE id=?", (dept_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'message': '部门删除成功'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400

@departments_bp.route('/export', methods=['GET'])
@permission_required('departments', 'export')
def export_departments_excel():
    """按单机版格式导出部门及员工。"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from io import BytesIO

    conn = get_connection()
    c = conn.cursor()
    c.execute('SELECT id, name, manager, contact, address FROM departments ORDER BY id')
    departments = c.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '部门员工清单'
    ws.append(DEPARTMENT_EMPLOYEE_HEADERS)

    unique_employees = set()
    data_row_count = 1
    for dept_id, name, manager, contact, address in departments:
        if manager:
            unique_employees.add(manager)
        c.execute('''SELECT name, position, contact FROM employees
                     WHERE department_id=? ORDER BY sort_order, id''', (dept_id,))
        employees = c.fetchall()
        data_row_count += max(1, len(employees))
        if not employees:
            ws.append([name, manager, contact, address, '', '', '', ''])
            continue

        for index, (emp_name, position, emp_contact) in enumerate(employees, 1):
            unique_employees.add(emp_name)
            dept_values = [name, manager, contact, address] if index == 1 else ['', '', '', '']
            ws.append(dept_values + [f'{index:02d}', emp_name, position, emp_contact])

    conn.close()

    ws.append([])
    ws.append(['统计信息'])
    ws.append(['总部门数', len(departments)])
    ws.append(['总员工数', len(unique_employees)])
    ws.append([])
    ws.append([
        '导出说明：\n'
        '1. 同一部门的员工需要连续填写，部门信息只在第一行填写。\n'
        '2. 部门内员工序号从01开始连续编号。\n'
        '3. 如果只有部门没有员工，员工信息列可以留空。\n'
        '4. 导入时会根据部门名称自动分组处理。\n'
        '5. 请勿修改标题行的列名和顺序。'
    ])
    ws.append([])
    ws.append(['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

    for index, width in enumerate([20, 15, 15, 50, 15, 15, 15, 15], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    for cell in ws[1]:
        cell.font = Font(bold=True, name='微软雅黑', size=11)
        cell.alignment = Alignment(horizontal='center')

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=data_row_count, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"部门员工列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@departments_bp.route('/import-template', methods=['GET'])
@permission_required('departments', 'import')
def download_department_import_template():
    """下载与单机版 export_template 一致的部门员工导入模板。"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from io import BytesIO
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '部门员工清单'
    ws.append(DEPARTMENT_EMPLOYEE_HEADERS)
    examples = [
        ['第一示例部门（导入前请删除示例行）', '示例负责人甲', 'DEMO-PHONE-001', '示例市第一示例地址', '01', '示例员工甲', '负责人', 'DEMO-MOBILE-001'],
        ['', '', '', '', '02', '示例员工乙', '员工', 'DEMO-MOBILE-002'],
        ['第二示例部门（这是新增部门的写法）', '示例负责人乙', 'DEMO-PHONE-002', '示例市第二示例地址', '01', '示例员工丙', '负责人', 'DEMO-MOBILE-003'],
        ['', '', '', '', '02', '示例员工丁', '员工', 'DEMO-MOBILE-004']
    ]
    for row in examples: ws.append(row)
    for index, width in enumerate([20, 15, 15, 50, 15, 15, 15, 15], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    thin = Side(style='thin'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=8):
        for cell in row: cell.border = border; cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]: cell.font = Font(bold=True, name='微软雅黑', size=11)
    ws.append([])
    ws.append(['导出说明：\n1. 同一部门的员工需要连续填写，部门信息只在第一行填写。\n2. 部门内员工序号从01开始连续编号。\n3. 如果只有部门没有员工，员工信息列可以留空。\n4. 导入时会根据部门名称自动分组处理。\n5. 请勿修改标题行的列名和顺序。'])
    ws.row_dimensions[7].height = 80
    output = BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name='部门员工导入模板.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@departments_bp.route('/import', methods=['POST'])
@permission_required('departments', 'import')
def import_departments_excel():
    """导入单机版部门员工格式，同时兼容旧模板和纯部门列表。"""
    from io import BytesIO
    from openpyxl import load_workbook

    upload = request.files.get('file')
    if not upload:
        return jsonify({'error': '请上传Excel文件'}), 400

    try:
        wb = load_workbook(BytesIO(upload.read()), data_only=True)
    except Exception as exc:
        return jsonify({'error': f'无法读取Excel文件：{exc}'}), 400

    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
    is_template = headers == DEPARTMENT_EMPLOYEE_HEADERS
    is_old_template = headers == OLD_DEPARTMENT_EMPLOYEE_HEADERS

    normalized = [header.lower() for header in headers]
    name_index = next((i for i, value in enumerate(normalized) if '部门名称' in value or value == '名称'), None)
    manager_index = next((i for i, value in enumerate(normalized) if '负责人' in value or '经理' in value or 'manager' in value), None)
    contact_index = next((i for i, value in enumerate(normalized) if '联系' in value or 'contact' in value or 'phone' in value), None)
    address_index = next((i for i, value in enumerate(normalized) if '地址' in value or 'address' in value or 'location' in value), None)
    is_department_list = not (is_template or is_old_template) and name_index is not None and sum(
        value is not None for value in (manager_index, contact_index, address_index)
    ) >= 2

    if not (is_template or is_old_template or is_department_list):
        return jsonify({
            'error': '无法识别Excel格式，请使用单机版导出的“部门员工清单”或标准部门列表。',
            'headers': headers
        }), 400

    conn = get_connection()
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    dept_imported = 0
    employee_imported = 0
    duplicate_departments = 0
    duplicate_employees = 0
    current_dept_id = None

    def text(value):
        return str(value).strip() if value is not None else ''

    try:
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(value is not None and text(value) for value in row):
                continue

            first_value = text(row[0]) if row else ''
            if first_value in ('统计信息', '总部门数', '总员工数', '导出时间') or first_value.startswith('导出说明：'):
                continue

            if is_template or is_old_template:
                values = list(row[:8]) + [None] * max(0, 8 - len(row))
                if is_old_template:
                    dept_name, dept_address, dept_contact, dept_manager, emp_seq, emp_name, emp_position, emp_contact = values[:8]
                else:
                    dept_name, dept_manager, dept_contact, dept_address, emp_seq, emp_name, emp_position, emp_contact = values[:8]

                dept_name, dept_manager = text(dept_name), text(dept_manager)
                dept_contact, dept_address = text(dept_contact), text(dept_address)
                emp_name, emp_position, emp_contact = text(emp_name), text(emp_position), text(emp_contact)

                if dept_name:
                    c.execute('SELECT id FROM departments WHERE name=?', (dept_name,))
                    existing = c.fetchone()
                    if existing:
                        current_dept_id = existing[0]
                        duplicate_departments += 1
                    else:
                        c.execute('''INSERT INTO departments (name, manager, contact, address, created_at)
                                     VALUES (?, ?, ?, ?, ?)''',
                                  (dept_name, dept_manager, dept_contact, dept_address, now))
                        current_dept_id = c.lastrowid
                        dept_imported += 1

                if emp_name and current_dept_id:
                    c.execute('SELECT id FROM employees WHERE name=? AND department_id=?',
                              (emp_name, current_dept_id))
                    if c.fetchone():
                        duplicate_employees += 1
                    else:
                        try:
                            sort_order = int(float(emp_seq)) if emp_seq not in (None, '') else 1
                        except (TypeError, ValueError):
                            sort_order = 1
                        c.execute('''INSERT INTO employees
                                     (name, position, department_id, contact, sort_order, created_at)
                                     VALUES (?, ?, ?, ?, ?, ?)''',
                                  (emp_name, emp_position, current_dept_id, emp_contact, sort_order, now))
                        employee_imported += 1
            else:
                dept_name = text(row[name_index]) if name_index < len(row) else ''
                if not dept_name:
                    continue
                dept_manager = text(row[manager_index]) if manager_index is not None and manager_index < len(row) else ''
                dept_contact = text(row[contact_index]) if contact_index is not None and contact_index < len(row) else ''
                dept_address = text(row[address_index]) if address_index is not None and address_index < len(row) else ''
                c.execute('SELECT id FROM departments WHERE name=?', (dept_name,))
                if c.fetchone():
                    duplicate_departments += 1
                    continue
                c.execute('''INSERT INTO departments (name, manager, contact, address, created_at)
                             VALUES (?, ?, ?, ?, ?)''',
                          (dept_name, dept_manager, dept_contact, dept_address, now))
                dept_imported += 1

        conn.commit()
    except Exception as exc:
        conn.rollback()
        conn.close()
        return jsonify({'error': f'第{row_number}行导入失败：{exc}'}), 400

    conn.close()
    return jsonify({
        'departments_imported': dept_imported,
        'employees_imported': employee_imported,
        'duplicate_departments': duplicate_departments,
        'duplicate_employees': duplicate_employees
    })
