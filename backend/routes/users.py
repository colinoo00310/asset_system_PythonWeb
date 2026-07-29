"""
用户管理路由
"""

from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from config.database import get_connection, hash_password
from routes.auth import role_required

users_bp = Blueprint('users', __name__)

@users_bp.route('/import-template', methods=['GET'])
@role_required('admin')
def download_user_import_template():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '用户导入模板'
    ws.append(['用户名', '密码', '角色', '姓名', '所属部门'])
    ws.append(['example', '请修改密码', '员工', '示例用户（导入前请删除本行）', '默认部门'])
    ws.append([]); ws.append(['填写说明', '角色可填写：管理员、科长、员工；所属部门必须与系统部门名称完全一致。'])
    for cell in ws[1]: cell.font = Font(bold=True, name='微软雅黑'); cell.alignment = Alignment(horizontal='center')
    for column_index in range(1, 6):
        values = [str(ws.cell(row=row, column=column_index).value or '') for row in range(1, ws.max_row + 1)]
        ws.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = min(max(max(map(len, values)) + 4, 18), 42)
    output = BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name='用户导入模板.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@users_bp.route('/import', methods=['POST'])
@role_required('admin')
def import_users_excel():
    from openpyxl import load_workbook
    from io import BytesIO
    upload = request.files.get('file')
    if not upload: return jsonify({'error': '请上传Excel文件'}), 400
    try: wb = load_workbook(BytesIO(upload.read()), data_only=True)
    except Exception as exc: return jsonify({'error': f'无法读取Excel文件：{exc}'}), 400
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    if headers[:5] != ['用户名', '密码', '角色', '姓名', '所属部门']:
        return jsonify({'error': '模板格式不正确，请下载用户导入模板后填写'}), 400
    role_map = {'管理员': 'admin', '科长': 'section_chief', '员工': 'staff', 'admin': 'admin', 'section_chief': 'section_chief', 'staff': 'staff'}
    conn = get_connection(); c = conn.cursor(); imported = 0; duplicate = 0; errors = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row): continue
        username, password, role_value, full_name, dept_name = [(str(v).strip() if v is not None else '') for v in (list(row[:5]) + [''] * 5)[:5]]
        if not username or not password: errors.append(f'第{row_no}行：用户名和密码不能为空'); continue
        role = role_map.get(role_value)
        if not role: errors.append(f'第{row_no}行：角色无效'); continue
        c.execute('SELECT id FROM users WHERE username=?', (username,))
        if c.fetchone(): duplicate += 1; continue
        department_id = None
        if dept_name:
            c.execute('SELECT id FROM departments WHERE name=?', (dept_name,)); dept = c.fetchone()
            if not dept: errors.append(f'第{row_no}行：所属部门“{dept_name}”不存在'); continue
            department_id = dept[0]
        if role != 'admin' and department_id is None: errors.append(f'第{row_no}行：科长和员工必须填写所属部门'); continue
        c.execute('INSERT INTO users (username,password,role,full_name,department_id) VALUES (?,?,?,?,?)',
                  (username, hash_password(password), role, full_name, department_id)); imported += 1
    conn.commit(); conn.close()
    return jsonify({'imported': imported, 'duplicate': duplicate, 'errors': errors[:50]})

@users_bp.route('/', methods=['GET'])
@role_required('admin')
def get_users():
    """获取用户列表"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("""SELECT u.id, u.username, u.role, u.full_name, u.department_id, d.name as department_name
                 FROM users u
                 LEFT JOIN departments d ON u.department_id = d.id
                 ORDER BY 
                     CASE u.role
                         WHEN 'admin' THEN 0
                         WHEN 'section_chief' THEN 1
                         WHEN 'staff' THEN 2
                         ELSE 3
                     END, u.id""")
    users = c.fetchall()
    conn.close()
    
    result = []
    for user in users:
        result.append({
            'id': user[0],
            'username': user[1],
            'role': user[2],
            'full_name': user[3],
            'department_id': user[4],
            'department_name': user[5]
        })
    
    return jsonify({'users': result})

@users_bp.route('/<int:user_id>', methods=['GET'])
@role_required('admin')
def get_user(user_id):
    """获取单个用户"""
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("""SELECT u.id, u.username, u.role, u.full_name, u.department_id, d.name as department_name
                 FROM users u
                 LEFT JOIN departments d ON u.department_id = d.id
                 WHERE u.id=?""", (user_id,))
    user = c.fetchone()
    conn.close()
    
    if not user:
        return jsonify({'error': '用户不存在'}), 404
    
    return jsonify({
        'id': user[0],
        'username': user[1],
        'role': user[2],
        'full_name': user[3],
        'department_id': user[4],
        'department_name': user[5]
    })

@users_bp.route('/', methods=['POST'])
@role_required('admin')
def create_user():
    """创建用户"""
    current_user = request.current_user
    
    if current_user['role'] != 'admin':
        return jsonify({'error': '只有管理员可以创建用户'}), 403
    
    data = request.get_json()
    
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    role = data.get('role', 'staff')
    full_name = data.get('full_name', '').strip()
    department_id = data.get('department_id')
    
    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400
    
    if role not in ['admin', 'section_chief', 'staff']:
        return jsonify({'error': '无效的角色'}), 400
    
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("SELECT id FROM users WHERE username=?", (username,))
    if c.fetchone():
        conn.close()
        return jsonify({'error': '用户名已存在'}), 400
    
    hashed_pwd = hash_password(password)
    
    try:
        c.execute("""INSERT INTO users (username, password, role, full_name, department_id)
                     VALUES (?, ?, ?, ?, ?)""",
                  (username, hashed_pwd, role, full_name, department_id))
        user_id = c.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'message': '用户创建成功', 'id': user_id}), 201
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400

@users_bp.route('/<int:user_id>', methods=['PUT'])
@role_required('admin')
def update_user(user_id):
    """更新用户"""
    current_user = request.current_user
    
    if current_user['role'] != 'admin':
        return jsonify({'error': '只有管理员可以修改用户'}), 403
    
    data = request.get_json()
    conn = get_connection()
    c = conn.cursor()
    
    c.execute("SELECT id FROM users WHERE id=?", (user_id,))
    if not c.fetchone():
        conn.close()
        return jsonify({'error': '用户不存在'}), 404
    
    update_fields = []
    params = []
    
    if 'password' in data and data['password']:
        update_fields.append("password = ?")
        params.append(hash_password(data['password']))
    
    if 'role' in data:
        if data['role'] not in ['admin', 'section_chief', 'staff']:
            conn.close()
            return jsonify({'error': '无效的角色'}), 400
        update_fields.append("role = ?")
        params.append(data['role'])
    
    if 'full_name' in data:
        update_fields.append("full_name = ?")
        params.append(data['full_name'])
    
    if 'department_id' in data:
        update_fields.append("department_id = ?")
        params.append(data['department_id'])
    
    if not update_fields:
        conn.close()
        return jsonify({'error': '没有要更新的字段'}), 400
    
    params.append(user_id)
    
    try:
        c.execute(f"UPDATE users SET {', '.join(update_fields)} WHERE id=?", params)
        conn.commit()
        conn.close()
        
        return jsonify({'message': '用户更新成功'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400

@users_bp.route('/<int:user_id>', methods=['DELETE'])
@role_required('admin')
def delete_user(user_id):
    """删除用户"""
    current_user = request.current_user
    
    if current_user['role'] != 'admin':
        return jsonify({'error': '只有管理员可以删除用户'}), 403
    
    if current_user['user_id'] == user_id:
        return jsonify({'error': '不能删除当前登录的用户'}), 400
    
    conn = get_connection()
    c = conn.cursor()
    
    try:
        c.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'message': '用户删除成功'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 400
