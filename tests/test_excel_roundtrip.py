import io
import os
import re
import shutil
import tempfile
import unittest
import sqlite3

from openpyxl import load_workbook


ASSET_HEADERS = [
    "ID", "名称", "管理方式", "资产分类", "编号", "数量",
    "购置日期", "市场价值", "责任人", "地址", "状态",
    "租赁开始日期", "租赁结束日期", "承租方", "承租方联系方式",
    "承租方性质", "承租方用途", "租金金额", "租金交付方式", "公开招拍租情况",
    "产证情况", "产权单位", "建筑面积（平方米）",
    "托管合同类型", "托管合同金额", "托管合同相对方",
    "托管合同编号", "托管合同开始日期", "托管合同结束日期", "托管签署日期",
    "托管是否归档", "录入人", "所属部门", "备注"
]

DEPARTMENT_HEADERS = [
    "部门名称", "负责人", "联系方式", "部门地址",
    "部门内员工序号", "员工姓名", "员工职位", "员工联系方式"
]


class ExcelRoundTripTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.mkdtemp(prefix='asset-system-excel-')
        cls.db_path = os.path.join(cls.temp_dir, 'assets.db')
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        os.environ['ASSET_DB_PATH'] = cls.db_path
        os.environ['ASSET_IMAGES_DIR'] = os.path.join(cls.temp_dir, 'images')

        from backend.app import create_app
        cls.client = create_app().test_client()
        captcha = cls.client.get('/api/auth/captcha').get_json()
        answer = sum(map(int, re.findall(r'\d+', captcha['question'])))
        response = cls.client.post('/api/auth/login', json={
            'username': 'admin',
            'password': 'admin123',
            'captcha_id': captcha['captcha_id'],
            'captcha': answer,
        })
        cls.headers = {'Authorization': f"Bearer {response.get_json()['token']}"}

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.temp_dir, ignore_errors=True)

    def test_asset_export_and_import_roundtrip(self):
        exported = self.client.get('/api/assets/export', headers=self.headers)
        self.assertEqual(exported.status_code, 200)
        workbook = load_workbook(io.BytesIO(exported.data))
        sheet = workbook.active
        self.assertEqual(sheet.title, '资产清单')
        self.assertEqual([cell.value for cell in sheet[1]], ASSET_HEADERS)
        self.assertEqual(sheet.freeze_panes, 'A2')
        self.assertEqual(sheet.row_dimensions[1].height, 30)

        sheet.cell(2, 1).value = 999999
        sheet.cell(2, 2).value = 'Excel往返测试资产'
        sheet.cell(2, 5).value = 'CODEX-ROUNDTRIP-ASSET'
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        imported = self.client.post(
            '/api/assets/import',
            headers=self.headers,
            data={'file': (payload, '资产清单.xlsx')},
            content_type='multipart/form-data',
        )
        self.assertEqual(imported.status_code, 200, imported.get_data(as_text=True))
        self.assertEqual(imported.get_json()['imported'], 1)

    def test_legacy_rental_assets_have_display_status(self):
        response = self.client.get('/api/assets/?per_page=500', headers=self.headers)
        self.assertEqual(response.status_code, 200)
        rental_assets = [
            asset for asset in response.get_json()['assets']
            if asset.get('management_type') == '租赁管理'
        ]
        self.assertTrue(rental_assets)
        self.assertTrue(all(asset.get('lease_status') for asset in rental_assets))

    def test_asset_list_uses_desktop_ascending_order(self):
        response = self.client.get('/api/assets/?page=1&per_page=20', headers=self.headers)
        self.assertEqual(response.status_code, 200)
        asset_ids = [asset['id'] for asset in response.get_json()['assets']]
        self.assertEqual(asset_ids, sorted(asset_ids))

    def test_import_templates_download(self):
        asset_template = self.client.get('/api/assets/import-template', headers=self.headers)
        user_template = self.client.get('/api/users/import-template', headers=self.headers)
        self.assertEqual(asset_template.status_code, 200)
        self.assertEqual(user_template.status_code, 200)
        self.assertEqual(load_workbook(io.BytesIO(asset_template.data)).active.max_column, 34)
        asset_sheet = load_workbook(io.BytesIO(asset_template.data)).active
        user_sheet = load_workbook(io.BytesIO(user_template.data)).active
        self.assertGreater(asset_sheet.column_dimensions['B'].width, 12)
        self.assertEqual(user_sheet.max_column, 5)
        self.assertGreater(user_sheet.column_dimensions['E'].width, 14)
        department_template = self.client.get('/api/departments/import-template', headers=self.headers)
        self.assertEqual(department_template.status_code, 200)
        department_sheet = load_workbook(io.BytesIO(department_template.data)).active
        self.assertEqual(department_sheet.max_column, 8)
        self.assertIn('第二示例部门', department_sheet['A4'].value)

        imported = self.client.post('/api/assets/import', headers=self.headers,
            data={'file': (io.BytesIO(asset_template.data), '资产导入模板.xlsx')}, content_type='multipart/form-data')
        self.assertEqual(imported.status_code, 200)
        self.assertEqual(imported.get_json()['imported'], 1)

    def test_chinese_named_png_upload(self):
        assets = self.client.get('/api/assets/?per_page=1', headers=self.headers).get_json()['assets']
        response = self.client.post(f"/api/assets/{assets[0]['id']}/images", headers=self.headers,
            data={'images': (io.BytesIO(b'fake-png-test'), '资产照片.png')}, content_type='multipart/form-data')
        self.assertEqual(response.status_code, 200, response.get_json())
        first_path = response.get_json()['images'][0]
        appended = self.client.post(f"/api/assets/{assets[0]['id']}/images", headers=self.headers,
            data={'keep_images': first_path, 'images': (io.BytesIO(b'fake-jpg-test'), '新增照片.jpg')},
            content_type='multipart/form-data')
        self.assertEqual(appended.status_code, 200, appended.get_json())
        self.assertEqual(len(appended.get_json()['images']), 2)
        removed = self.client.post(f"/api/assets/{assets[0]['id']}/images", headers=self.headers,
            data={'keep_images': first_path}, content_type='multipart/form-data')
        self.assertEqual(removed.status_code, 200, removed.get_json())
        self.assertEqual(removed.get_json()['images'], [first_path])

    def test_department_user_only_sees_own_assets(self):
        from backend.config.database import hash_password
        conn = sqlite3.connect(self.db_path); cursor = conn.cursor()
        cursor.execute("INSERT INTO departments(name,address,manager,created_at) VALUES (?,?,?,?)",
                       ('权限测试部门', '测试地址', '测试负责人', '2026-01-01'))
        department_id = cursor.lastrowid
        cursor.execute("INSERT INTO users(username,password,role,full_name,department_id) VALUES (?,?,?,?,?)",
                       ('scope_test', hash_password('scope123'), 'staff', '权限测试用户', department_id))
        cursor.execute("INSERT INTO assets(name,category,management_type,asset_number,responsible_person,location,department_id) VALUES (?,?,?,?,?,?,?)",
                       ('本部门测试资产', '其他资产', '自主管理', 'SCOPE-OWN', '测试人', '测试地址', department_id))
        conn.commit(); conn.close()
        captcha = self.client.get('/api/auth/captcha').get_json()
        answer = sum(map(int, re.findall(r'\d+', captcha['question'])))
        login = self.client.post('/api/auth/login', json={
            'username': 'scope_test', 'password': 'scope123',
            'captcha_id': captcha['captcha_id'], 'captcha': answer,
        }).get_json()
        chief_headers = {'Authorization': f"Bearer {login['token']}"}
        me = self.client.get('/api/auth/me', headers=chief_headers).get_json()
        assets = self.client.get('/api/assets/?per_page=500', headers=chief_headers).get_json()['assets']
        self.assertTrue(assets)
        self.assertTrue(all(asset['department_id'] == me['department_id'] for asset in assets))

    def test_department_export_and_import_roundtrip(self):
        exported = self.client.get('/api/departments/export', headers=self.headers)
        self.assertEqual(exported.status_code, 200)
        workbook = load_workbook(io.BytesIO(exported.data))
        sheet = workbook.active
        self.assertEqual(sheet.title, '部门员工清单')
        self.assertEqual([cell.value for cell in sheet[1]], DEPARTMENT_HEADERS)

        sheet.cell(2, 1).value = 'Excel往返测试部门'
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        imported = self.client.post(
            '/api/departments/import',
            headers=self.headers,
            data={'file': (payload, '部门员工列表.xlsx')},
            content_type='multipart/form-data',
        )
        self.assertEqual(imported.status_code, 200, imported.get_data(as_text=True))
        self.assertGreaterEqual(imported.get_json()['departments_imported'], 1)


if __name__ == '__main__':
    unittest.main()
